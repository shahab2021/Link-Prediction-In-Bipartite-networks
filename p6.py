#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/env python
# coding: utf-8

# In[77]:


import openpyxl
import math
class Link_Prediction:
    
    def __init__(self):
        self.c1_dict = {}
        self.c2_dict = {}
        self.output_score = []
        self.output_dict = {}
        
    def read_xlsx(self,path_read):
        wb_obj = openpyxl.load_workbook(path_read)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row #getting the number of rows
        c1_dict = {} #defining a dictionary
        c2_dict = {} #defining a dictionary
        for r in range(max_row):
            r = r+1 #skipping the first row
            if r > 1: 
                c1 = str(sheet_obj.cell(r,1).value) #extracting first column data
                c2 = str(sheet_obj.cell(r,2).value) #extracting second column data    
                if c1 in c1_dict: #checking the existence of c1 in c1_dict's keys
                    if c2 not in c1_dict[c1].split(","): #checking the absence of c2 in c1's values
                        c1_dict[c1] += "," + c2 #adding c2 to c1's values
                else:
                    c1_dict[c1] = c2 #setting c1 as a key and c2 as a value to c1_dict

                if c2 in c2_dict: #checking the existence of c2 in c2_dict's keys
                    if c1 not in c2_dict[c2].split(","): #checking the absence of c1 in c2's values
                        c2_dict[c2] += "," + c1 #adding c1 to c2's values
                else:
                    c2_dict[c2] = c1 #setting c2 as a key and c1 as a value to c2_dict
        self.c1_dict = c1_dict
        self.c2_dict = c2_dict
        
    def PA(self):
        c1_dict = self.c1_dict
        c2_dict = self.c2_dict
        c1_keys = list(c1_dict.keys()) #getting all the keys of c1_dict as a list
        c2_keys = list(c2_dict.keys()) #getting all the keys of c2_dict as a list
        result = [] #defining a list
        for c1 in c1_keys:
            c2_list = c1_dict[c1].split(",") # getting the values of c1 key
            for c2 in c2_keys:
                if c2 not in c2_list: #checking the absence of c2 in c2_list
                    lc2 = len(c2_dict[c2].split(",")) #getting the number of c2'values
                    lc1 = len(c1_dict[c1].split(","))#getting the number of c1'values
                    result.append(c1+","+c2+","+str(lc1*lc2)) #calculating the PA scores
        self.sort(result)
    
    def CN(self):
        c1_dict = self.c1_dict
        c2_dict = self.c2_dict
        c1_keys = list(c1_dict.keys()) #getting all the keys of c1_dict as a list
        c2_keys = list(c2_dict.keys()) #getting all the keys of c2_dict as a list
        result = [] #defining a list
        for c1 in c1_keys:
            c2_list = c1_dict[c1].split(",") # getting the values of c1 key
            for c2 in c2_keys:
                if c2 not in c2_list: #checking the absence of c2 in c2_list
                    cc1_list = c2_dict[c2].split(",") #from here onwards: calculating r'(c2)
                    cc2_set = set() #there isn't any duplicate elements in a set structure
                    for cc1 in cc1_list:
                        cc2_set = cc2_set.union(set(c1_dict[cc1].split(",")))
                    ccc2_set = set(c1_dict[c1].split(','))
                    common_c2 = cc2_set.intersection(ccc2_set) #common_c2 indicates the set of common members
                    result.append(c1+","+c2+","+str(len(common_c2)))
        self.sort(result)
        
    def JC(self):
        c1_dict = self.c1_dict
        c2_dict = self.c2_dict
        c1_keys = list(c1_dict.keys()) #getting all the keys of c1_dict as a list
        c2_keys = list(c2_dict.keys()) #getting all the keys of c2_dict as a list
        result = [] #defining a list
        for c1 in c1_keys:
            c2_list = c1_dict[c1].split(",") # getting the values of c1 key
            for c2 in c2_keys:
                if c2 not in c2_list: #checking the absence of c2 in c2_list
                    cc1_list = c2_dict[c2].split(",") #from here onwards: calculating r'(c2)
                    cc2_set = set() #there isn't any duplicate elements in a set structure
                    for cc1 in cc1_list:
                        cc2_set = cc2_set.union(set(c1_dict[cc1].split(",")))
                    ccc2_set = set(c1_dict[c1].split(','))
                    common_c2 = cc2_set.intersection(ccc2_set) #common_c2 indicates the set of common members
                    cc2_set = cc2_set.union(ccc2_set)
                    result.append(c1+","+c2+","+str(len(common_c2)/len(cc2_set)))
        self.sort(result)
        
    def AA(self):
        c1_dict = self.c1_dict
        c2_dict = self.c2_dict
        c1_keys = list(c1_dict.keys()) #getting all the keys of c1_dict as a list
        c2_keys = list(c2_dict.keys()) #getting all the keys of c2_dict as a list
        result = [] #defining a list
        for c1 in c1_keys:
            c2_list = c1_dict[c1].split(",") # getting the values of c1 key
            for c2 in c2_keys:
                if c2 not in c2_list: #checking the absence of c2 in c2_list
                    cc1_list = c2_dict[c2].split(",") #from here onwards: calculating r'(c2)
                    cc2_set = set() #there isn't any duplicate elements in a set structure
                    for cc1 in cc1_list:
                        cc2_set = cc2_set.union(set(c1_dict[cc1].split(",")))
                    ccc2_set = set(c1_dict[c1].split(','))
                    common_c2 = cc2_set.intersection(ccc2_set) #common_c2 indicates the set of common members
                    r = 0.0
                    for cccc2 in common_c2:
                        r += 1/math.log(len(c2_dict[cccc2].split(","))) 
                    result.append(c1+","+c2+","+str(r))
        self.sort(result)
    
    def sort(self,result):
        th2_dict = {} #defining a dictionary
        for r in result:
            th = r.split(",") #splitting r to th list
            if float(th[2]) in th2_dict: #checking the existence of third member of th(score) in sort_dict
                th2_dict[float(th[2])] += "&" + r #adding r to th[2]'s values
            else:
                th2_dict[float(th[2])] = r #setting th[2] as a key and r as a value to th2_dict
        sort = list(th2_dict.keys()) #getting all the keys of th2_dict as a list
        sort.sort(reverse=True) #sorting the sort list
        self.output_score = sort
        self.output_dict = th2_dict
    
    def write(self,path_write):
        out_score = self.output_score # out_score is a list of scores
        out_dict = self.output_dict # out_dict is a dictionary, the key is score and the value is c1,c2,score
        wb = openpyxl.Workbook() 
        sheet = wb.active
        count = 1
        for os in out_score: # os = score
            l = out_dict[os].split("&") # out_dict[os] = c1,c2,score & c1,c2,score...
            for i in l:# i = c1,c2,score
                i_list = i.split(",") 
                sheet.cell(row= count,column = 1).value = i_list[0]
                sheet.cell(row= count,column = 2).value = i_list[1]
                sheet.cell(row= count,column = 3).value = i_list[2]
                count += 1
        wb.save(path_write)